import io
import re
import zipfile
from pathlib import Path
from datetime import datetime
from urllib.parse import quote
from html import escape
from typing import Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st

try:
    from pypdf import PdfReader
except Exception:
    try:
        from PyPDF2 import PdfReader
    except Exception:
        PdfReader = None

PAS_YELLOW = "#FFD400"
PAS_BLACK = "#0A0A0A"
PAS_DARK = "#171717"
PAS_GREY = "#F4F4F4"

st.set_page_config(page_title="PAS Vehicle Hire Matching", page_icon="pas_logo.png", layout="wide")

st.markdown(
    f"""
    <style>
    .stApp {{ background: #f5f5f5; color: #0A0A0A; }}
    section[data-testid="stSidebar"] {{
        background: {PAS_BLACK};
        color: white;
        padding-top: 1.45rem;
    }}
    section[data-testid="stSidebar"] * {{ color: white; }}
    section[data-testid="stSidebar"] img {{
        margin-top: 0.15rem;
        border-radius: 14px;
    }}
    .block-container {{
        padding-top: 1.4rem;
        padding-bottom: 2rem;
        max-width: 1500px;
    }}

    .pas-hero {{
        background: linear-gradient(135deg, {PAS_BLACK} 0%, #202020 70%, #7a6900 135%);
        border-radius: 18px;
        padding: 24px 28px;
        margin-bottom: 18px;
        box-shadow: 0 8px 25px rgba(0,0,0,0.12);
    }}
    .pas-title {{
        color: white;
        font-size: 32px;
        font-weight: 900;
        margin: 0;
        letter-spacing: -0.03em;
    }}
    .pas-subtitle {{
        color: {PAS_YELLOW};
        font-size: 14px;
        margin-top: 4px;
        font-weight: 800;
    }}

    .kpi-card {{
        background: white;
        border-radius: 18px;
        padding: 18px 20px;
        border: 1px solid #e8e8e8;
        box-shadow: 0 3px 12px rgba(0,0,0,0.05);
        min-height: 112px;
    }}
    .kpi-label {{
        color: #111;
        font-size: 14px;
        font-weight: 800;
        margin-bottom: 8px;
    }}
    .kpi-value {{
        color: {PAS_YELLOW};
        font-size: 36px;
        font-weight: 950;
        line-height: 1.05;
        text-shadow: 0 1px 0 #111;
    }}
    .kpi-sub {{
        color: #222;
        font-size: 13px;
        margin-top: 6px;
    }}

    .stButton > button, .stDownloadButton > button {{
        background: {PAS_YELLOW} !important;
        color: {PAS_BLACK} !important;
        border: 1px solid {PAS_BLACK} !important;
        border-radius: 12px !important;
        font-weight: 900 !important;
    }}

    /* Keep app helper text readable */
    .stCaption, div[data-testid="stCaptionContainer"], .stMarkdown p, .stInfo {{
        color: #0A0A0A !important;
    }}

    .pas-results-title {{
        color: #0A0A0A;
        font-size: 26px;
        font-weight: 950;
        margin: 22px 0 8px 0;
    }}
    .pas-unmatched-pill {{
        display: inline-flex;
        align-items: center;
        gap: 8px;
        background: {PAS_YELLOW};
        color: {PAS_BLACK};
        border: 1px solid #111;
        border-radius: 14px 14px 0 0;
        padding: 11px 18px;
        font-weight: 950;
        box-shadow: 0 3px 10px rgba(0,0,0,0.08);
        margin-top: 4px;
    }}

    .pas-table-wrap {{
        background: white;
        border: 1px solid #d9d9d9;
        border-radius: 0 16px 16px 16px;
        overflow: auto;
        box-shadow: 0 4px 18px rgba(0,0,0,0.07);
        margin-bottom: 18px;
    }}
    table.pas-table {{
        width: 100%;
        border-collapse: collapse;
        font-size: 13px;
        color: #0A0A0A;
        background: white;
    }}
    table.pas-table thead th {{
        background: {PAS_YELLOW};
        color: {PAS_BLACK};
        font-weight: 950;
        text-align: left;
        padding: 11px 12px;
        border: 1px solid #c7a900;
        white-space: nowrap;
    }}
    table.pas-table tbody td {{
        background: white;
        color: #0A0A0A;
        padding: 9px 12px;
        border: 1px solid #e3e3e3;
        vertical-align: top;
    }}
    table.pas-table tbody tr:nth-child(even) td {{
        background: #fbfbfb;
    }}
    table.pas-table a {{
        color: #006fd6 !important;
        font-weight: 800;
        text-decoration: none;
    }}
    table.pas-table a:hover {{
        text-decoration: underline;
    }}
    .pas-note {{
        color: #0A0A0A;
        font-size: 13px;
        margin: 8px 0 16px 0;
    }}
    .pas-support {{
        color: #0A0A0A;
        font-size: 14px;
        margin: 16px 0;
    }}
    .pas-support a {{
        color: #006fd6 !important;
        font-weight: 800;
    }}


    /* --- hard hide Streamlit's uploaded-file chip/list while keeping uploader button usable --- */
    div[data-testid="stFileUploader"] [data-testid="stFileUploaderFile"],
    div[data-testid="stFileUploader"] [data-testid="stFileUploaderFileName"],
    div[data-testid="stFileUploader"] [data-testid="stFileUploaderFileSize"],
    div[data-testid="stFileUploader"] ul,
    div[data-testid="stFileUploader"] div[role="list"],
    div[data-testid="stFileUploader"] div[role="listitem"] {{
        display: none !important;
        visibility: hidden !important;
        height: 0 !important;
        min-height: 0 !important;
        margin: 0 !important;
        padding: 0 !important;
        overflow: hidden !important;
    }}
    div[data-testid="stFileUploader"] div:has(button[title*="Remove"]),
    div[data-testid="stFileUploader"] div:has(button[aria-label*="Remove"]),
    div[data-testid="stFileUploader"] div:has(svg[data-testid="DeleteIcon"]) {{
        display: none !important;
    }}
    div[data-testid="stFileUploader"] section > div:not(:has(button)) {{
        display: none !important;
    }}
    div[data-testid="stFileUploader"] button {{
        background: #ffffff !important;
        color: #0A0A0A !important;
        border: 1px solid #d7dce3 !important;
        border-radius: 10px !important;
        font-weight: 900 !important;
        box-shadow: 0 2px 8px rgba(0,0,0,.06) !important;
    }}
    div[data-testid="stFileUploader"] button * {{ color:#0A0A0A !important; fill:#0A0A0A !important; stroke:#0A0A0A !important; }}

    </style>
    """,
    unsafe_allow_html=True,
)


st.markdown(
    """
    <style>
    /* Keep sidebar readable on black */
    section[data-testid="stSidebar"] .stMarkdown p,
    section[data-testid="stSidebar"] .stMarkdown li,
    section[data-testid="stSidebar"] .stMarkdown h1,
    section[data-testid="stSidebar"] .stMarkdown h2,
    section[data-testid="stSidebar"] .stMarkdown h3,
    section[data-testid="stSidebar"] .stMarkdown strong,
    section[data-testid="stSidebar"] .stMarkdown span {
        color: #ffffff !important;
    }

    /* Make upload icons visible on dark bars */
    div[data-testid="stFileUploader"] svg,
    div[data-testid="stFileUploader"] button svg,
    div[data-testid="stFileUploader"] [data-testid="stIconMaterial"] {
        color: #FFD400 !important;
        fill: #FFD400 !important;
        stroke: #FFD400 !important;
    }
    div[data-testid="stFileUploader"] section {
        background: #24242d !important;
        border: 1px solid #30303a !important;
        border-radius: 12px !important;
    }
    div[data-testid="stFileUploader"] button {
        color: white !important;
        border-color: #454552 !important;
        background: #111217 !important;
    }

    /* Results table: white body, yellow sticky header, 10-row scroll area */
    .pas-table-wrap {
        max-height: 510px !important;
        overflow-y: auto !important;
        overflow-x: auto !important;
    }
    .pas-table-wrap thead th {
        position: sticky;
        top: 0;
        z-index: 2;
    }
    .pas-note, .pas-support, .pas-support * {
        color: #0A0A0A !important;
    }

    /* Bottom chase animation: small, low, runs once */
    .pas-bottom-chase-wrap {
        position: fixed;
        left: calc(18rem + 22px);
        right: 42px;
        bottom: 12px;
        height: 58px;
        pointer-events: none;
        z-index: 1;
        overflow: hidden;
    }
    .pas-bottom-ground {
        position: absolute;
        left: 0;
        right: 0;
        bottom: 6px;
        border-bottom: 1px solid rgba(0,0,0,0.11);
    }
    .pas-chase-pack {
        position: absolute;
        bottom: 8px;
        left: -150px;
        width: 150px;
        height: 48px;
        animation: pas-chase-run 13s linear 1 forwards;
    }
    @keyframes pas-chase-run {
        0% { transform: translateX(-120px); opacity: 0; }
        8% { opacity: 1; }
        88% { opacity: 1; }
        100% { transform: translateX(calc(100vw - 90px)); opacity: 0; }
    }
    .pas-truck-mini {
        position: absolute;
        left: 0;
        bottom: 5px;
        width: 54px;
        height: 30px;
        filter: drop-shadow(0 1px 1px rgba(0,0,0,.22));
    }
    .pas-truck-bed {
        position: absolute;
        left: 0;
        top: 5px;
        width: 34px;
        height: 19px;
        background: #FFD400;
        border: 3px solid #0A0A0A;
        border-radius: 4px 2px 3px 5px;
        transform: skewX(-10deg);
    }
    .pas-truck-logo {
        position: absolute;
        left: 7px;
        top: 9px;
        font-size: 9px;
        font-weight: 950;
        color: #0A0A0A;
        line-height: 1;
        z-index: 3;
    }
    .pas-truck-cab {
        position: absolute;
        left: 30px;
        top: 7px;
        width: 19px;
        height: 18px;
        background: #FFD400;
        border: 3px solid #0A0A0A;
        border-radius: 3px 5px 3px 2px;
        z-index: 2;
    }
    .pas-truck-window {
        position: absolute;
        left: 34px;
        top: 10px;
        width: 7px;
        height: 7px;
        background: #a8d8e8;
        border: 2px solid #0A0A0A;
        border-radius: 2px;
        z-index: 4;
    }
    .pas-truck-nose {
        position: absolute;
        left: 47px;
        top: 17px;
        width: 8px;
        height: 8px;
        background: #FFD400;
        border: 3px solid #0A0A0A;
        border-left: none;
        border-radius: 0 3px 3px 0;
    }
    .pas-wheel {
        position: absolute;
        bottom: 0;
        width: 9px;
        height: 9px;
        background: #0A0A0A;
        border: 2px solid #222;
        border-radius: 50%;
        animation: pas-wheel-spin .32s linear infinite;
        z-index: 5;
    }
    .pas-wheel::after {
        content: "";
        position: absolute;
        inset: 2px;
        background: #FFD400;
        border-radius: 50%;
    }
    .pas-wheel.back { left: 13px; }
    .pas-wheel.front { left: 41px; }
    @keyframes pas-wheel-spin { to { transform: rotate(360deg); } }

    .pas-speed-lines { position: absolute; left: -30px; top: 17px; width: 24px; height: 18px; }
    .pas-speed-lines span { display:block; height:2px; background:#b9b9b9; margin:4px 0; border-radius:2px; animation: pas-flicker .55s linear infinite; }
    .pas-speed-lines span:nth-child(2) { width: 16px; margin-left: 8px; }
    .pas-speed-lines span:nth-child(3) { width: 11px; margin-left: 13px; }
    @keyframes pas-flicker { 50% { opacity:.25; transform: translateX(-5px); } }

    .pas-dust { position:absolute; left:-5px; bottom:0; width:34px; height:14px; opacity:.75; }
    .pas-dust span { position:absolute; bottom:0; background:#dac6a9; border-radius:50%; animation: pas-dust 1s linear infinite; }
    .pas-dust span:nth-child(1) { width:12px; height:6px; left:0; }
    .pas-dust span:nth-child(2) { width:16px; height:7px; left:10px; animation-delay:.2s; }
    .pas-dust span:nth-child(3) { width:11px; height:5px; left:23px; animation-delay:.4s; }
    @keyframes pas-dust { 50% { transform: translateX(-8px) scale(1.15); opacity:.4; } }

    .pas-stickman {
        position: absolute;
        left: 92px;
        bottom: 5px;
        width: 28px;
        height: 34px;
        animation: pas-runner-bob .35s ease-in-out infinite alternate;
    }
    @keyframes pas-runner-bob { from { transform: translateY(1px); } to { transform: translateY(-2px); } }
    .pas-stick-head {
        position:absolute;
        top:0;
        left:11px;
        width:8px;
        height:8px;
        border:2px solid #111;
        border-radius:50%;
        background:white;
    }
    .pas-stick-body { position:absolute; left:15px; top:9px; width:2px; height:13px; background:#111; transform: rotate(12deg); transform-origin:top; }
    .pas-stick-arm-a, .pas-stick-arm-b, .pas-stick-leg-a, .pas-stick-leg-b { position:absolute; width:2px; height:12px; background:#111; transform-origin:top; border-radius:2px; }
    .pas-stick-arm-a { left:15px; top:11px; transform: rotate(58deg); animation: pas-arm-a .35s linear infinite alternate; }
    .pas-stick-arm-b { left:15px; top:11px; transform: rotate(-50deg); animation: pas-arm-b .35s linear infinite alternate; }
    .pas-stick-leg-a { left:16px; top:21px; height:14px; transform: rotate(48deg); animation: pas-leg-a .35s linear infinite alternate; }
    .pas-stick-leg-b { left:16px; top:21px; height:14px; transform: rotate(-42deg); animation: pas-leg-b .35s linear infinite alternate; }
    @keyframes pas-arm-a { to { transform: rotate(-45deg); } }
    @keyframes pas-arm-b { to { transform: rotate(55deg); } }
    @keyframes pas-leg-a { to { transform: rotate(-45deg); } }
    @keyframes pas-leg-b { to { transform: rotate(48deg); } }
    </style>
    """,
    unsafe_allow_html=True,
)


st.markdown(
    f"""
    <style>
    /* ===== PAS V2 target layout overrides: safe Streamlit-native controls ===== */
    .stApp {{ background: #f7f8fa !important; color: #0A0A0A !important; font-family: Inter, "Segoe UI", Arial, sans-serif; }}
    .block-container {{ max-width: 1580px !important; padding-top: 1.05rem !important; padding-left: 2rem !important; padding-right: 2rem !important; padding-bottom: 2rem !important; }}

    section[data-testid="stSidebar"] {{ background: linear-gradient(180deg, #050606 0%, #0b1015 100%) !important; border-right: 1px solid #161b22; }}
    section[data-testid="stSidebar"] > div:first-child {{ padding-top: 1.05rem !important; }}
    section[data-testid="stSidebar"] img {{ border-radius: 14px !important; box-shadow: 0 10px 24px rgba(0,0,0,.26); }}
    .pas-sidebar-title {{ color:#fff; font-size:18px; font-weight:950; line-height:1.15; text-align:center; margin: 20px 0 8px; }}
    .pas-yellow-line {{ width:72px; height:4px; background:{PAS_YELLOW}; border-radius:99px; margin: 0 auto 22px; }}
    .pas-sidebar-copy {{ color:#fff !important; font-size:14px; line-height:1.52; font-weight:650; margin-bottom:24px; }}
    .pas-sidebar-rule {{ border-top:1px solid rgba(255,255,255,.22); margin:22px 0; }}
    .pas-sidebar-heading {{ color:{PAS_YELLOW}; font-size:19px; font-weight:950; margin: 0 0 16px; }}
    .pas-nav-row {{ display:grid; grid-template-columns: 26px 1fr; gap:10px; align-items:start; margin: 15px 0; color:#fff; font-weight:750; line-height:1.25; font-size:14px; }}
    .pas-nav-icon svg {{ width:21px; height:21px; stroke:{PAS_YELLOW}; stroke-width:2.4; fill:none; stroke-linecap:round; stroke-linejoin:round; }}
    .pas-sidebar-footer {{ color:#fff; font-size:12px; font-weight:800; margin-top:28px; }}

    .pas-hero {{ display:flex; align-items:center; gap:16px; background: linear-gradient(100deg, #08090b 0%, #151718 70%, #c9aa00 130%) !important; border-radius: 16px !important; padding: 12px 22px !important; margin: 0 0 18px 0 !important; box-shadow: 0 9px 25px rgba(0,0,0,.13) !important; min-height:60px; }}
    .pas-hero-logo {{ width:37px; height:37px; border-radius:7px; background:{PAS_YELLOW}; color:#000; display:inline-flex; align-items:center; justify-content:center; font-weight:950; font-size:14px; letter-spacing:-1px; }}
    .pas-hero-text {{ color:#fff; font-size:18px; font-weight:950; letter-spacing:-.02em; }}
    .pas-hero-dot {{ color:#fff; opacity:.8; margin: 0 7px; }}
    .pas-hero-version {{ color:{PAS_YELLOW}; font-weight:950; }}

    .pas-upload-card {{ background:#fff; border:1px solid #e5e7eb; border-radius:18px; box-shadow:0 5px 18px rgba(15,23,42,.08); padding:16px 18px 14px; margin-bottom:14px; }}
    .pas-upload-title {{ color:#0A0A0A; font-size:16px; font-weight:950; margin-bottom:10px; }}
    div[data-testid="stFileUploader"] {{ margin:0 !important; }}
    div[data-testid="stFileUploader"] label {{ display:none !important; }}
    div[data-testid="stFileUploader"] section {{ background:#f4f6f8 !important; border:1px solid #dfe4ea !important; border-radius:11px !important; min-height:52px !important; padding:8px 10px !important; }}
    div[data-testid="stFileUploader"] section * {{ color:#0A0A0A !important; }}
    div[data-testid="stFileUploader"] button {{ background:#fff !important; color:#0A0A0A !important; border:1px solid #d7dce3 !important; border-radius:10px !important; font-weight:900 !important; box-shadow:0 2px 8px rgba(0,0,0,.06) !important; }}
    div[data-testid="stFileUploader"] svg {{ color:#0A0A0A !important; fill:currentColor !important; stroke:currentColor !important; }}
    div[data-testid="stFileUploader"] [data-testid="stFileUploaderFile"] {{ background:#fff !important; border:1px solid #dfe4ea !important; border-radius:10px !important; color:#0A0A0A !important; }}
    div[data-testid="stFileUploader"] small {{ color:#4b5563 !important; }}

    div.stButton > button[kind="secondary"], .stButton > button {{ min-height:52px !important; font-size:16px !important; box-shadow:0 6px 18px rgba(255,212,0,.25) !important; }}
    .stDownloadButton > button {{ min-height:62px !important; font-size:20px !important; box-shadow:0 6px 18px rgba(255,212,0,.25) !important; }}

    .kpi-card {{ background:#fff !important; border-radius:18px !important; border:1px solid #e4e7eb !important; box-shadow:0 5px 20px rgba(15,23,42,.08) !important; min-height:118px !important; padding:18px 22px !important; display:flex; align-items:center; gap:18px; }}
    .kpi-icon {{ width:64px; height:64px; border-radius:50%; background:#fff5bd; display:flex; align-items:center; justify-content:center; flex:none; }}
    .kpi-icon svg {{ width:35px; height:35px; stroke:#0A0A0A; stroke-width:2.5; fill:none; stroke-linecap:round; stroke-linejoin:round; }}
    .kpi-label {{ color:#111 !important; font-size:15px !important; font-weight:950 !important; margin:0 0 3px !important; }}
    .kpi-value {{ color:#e9b900 !important; font-size:42px !important; line-height:.98 !important; font-weight:950 !important; text-shadow:none !important; }}
    .kpi-sub {{ color:#374151 !important; font-size:14px !important; margin-top:6px !important; }}
    .kpi-unmatched .kpi-value {{ color:#e12626 !important; }}
    .kpi-matched .kpi-value {{ color:#16a34a !important; }}

    .pas-results-title {{ color:#0A0A0A !important; font-size:28px !important; font-weight:950 !important; margin: 22px 0 8px !important; }}
    .pas-unmatched-pill {{ background:{PAS_YELLOW} !important; color:#0A0A0A !important; border:0 !important; border-radius:14px 14px 0 0 !important; padding:13px 20px !important; font-size:18px; box-shadow:0 4px 14px rgba(0,0,0,.09); }}
    .pas-table-wrap {{ background:#fff !important; border:1px solid #e0e4e9 !important; border-radius:0 16px 16px 16px !important; max-height:430px !important; overflow:auto !important; box-shadow:0 7px 25px rgba(15,23,42,.10) !important; }}
    table.pas-table {{ font-size:14px !important; color:#0A0A0A !important; }}
    table.pas-table thead th {{ background:{PAS_YELLOW} !important; color:#0A0A0A !important; border:1px solid #e2ba00 !important; padding:12px 14px !important; font-weight:950 !important; position:sticky; top:0; z-index:5; }}
    table.pas-table tbody td {{ background:#fff !important; color:#0A0A0A !important; border:1px solid #e1e5eb !important; padding:10px 14px !important; }}
    table.pas-table tbody tr:nth-child(even) td {{ background:#fbfcfd !important; }}
    .pas-pdf-icon {{ display:inline-flex; align-items:center; justify-content:center; width:17px; height:20px; background:#e11d2e; color:#fff; border-radius:3px; font-size:9px; font-weight:950; margin-right:8px; vertical-align:middle; }}
    table.pas-table a {{ color:#006bd6 !important; font-weight:850 !important; }}
    table.pas-table .query-cell {{ min-width:120px; white-space:nowrap; }}
    .pas-note, .pas-support, .pas-support * {{ color:#0A0A0A !important; }}
    .pas-support {{ margin-top:22px !important; font-size:15px !important; }}
    .pas-support a {{ color:#006bd6 !important; font-weight:900 !important; margin-left:12px; }}

    /* --- uploader chip cleanup: hide Streamlit's ugly uploaded-file pill and use our own card --- */
    div[data-testid="stFileUploader"] [data-testid="stFileUploaderFile"] {{ display: none !important; }}
    div[data-testid="stFileUploaderDropzone"] {{ background: transparent !important; border: 0 !important; padding: 0 !important; min-height: 0 !important; }}
    div[data-testid="stFileUploaderDropzoneInstructions"] {{ display: none !important; }}
    div[data-testid="stFileUploader"] section {{ background: transparent !important; border: 0 !important; min-height: 0 !important; padding: 0 !important; }}
    div[data-testid="stFileUploader"] button {{
        background: #ffffff !important;
        color: #0A0A0A !important;
        border: 1px solid #d7dce3 !important;
        border-radius: 10px !important;
        font-weight: 900 !important;
        min-height: 44px !important;
        box-shadow: 0 2px 8px rgba(0,0,0,.06) !important;
    }}
    .pas-file-card {{
        display:flex; align-items:center; gap:14px;
        background:#f4f6f8; border:1px solid #dfe4ea; border-radius:12px;
        padding:11px 14px; min-height:54px; margin: 4px 0 12px;
    }}
    .pas-file-icon {{ width:32px; height:32px; border-radius:8px; display:flex; align-items:center; justify-content:center; color:#fff; font-weight:950; font-size:11px; box-shadow:0 2px 8px rgba(0,0,0,.12); flex:none; }}
    .pas-file-icon.excel {{ background:#118a3b; }}
    .pas-file-icon.pdf {{ background:#df1f2d; }}
    .pas-file-main {{ flex:1; min-width:0; }}
    .pas-file-name {{ color:#0A0A0A; font-weight:950; font-size:15px; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }}
    .pas-file-size {{ color:#4b5563; font-weight:650; font-size:13px; margin-top:2px; }}
    .pas-file-check {{ width:24px; height:24px; border-radius:50%; background:#108a37; color:white; display:flex; align-items:center; justify-content:center; font-size:15px; font-weight:950; flex:none; }}
    </style>
    """,
    unsafe_allow_html=True,
)

with st.sidebar:
    st.image("pas_logo.png", use_column_width=True)
    st.markdown(
        """
        <div class="pas-sidebar-title">PAS Vehicle Hire<br>Matching</div>
        <div class="pas-yellow-line"></div>
        <div class="pas-sidebar-copy">Upload the Vehicle spreadsheet and vehicle hire invoice PDF, then export annotated PDF and Excel report.</div>
        <div class="pas-sidebar-rule"></div>
        <div class="pas-sidebar-heading">Instructions</div>
        <div class="pas-nav-row"><span class="pas-nav-icon"><svg viewBox="0 0 24 24"><path d="M16 16l-4-4-4 4"/><path d="M12 12v9"/><path d="M20 16.6A5 5 0 0 0 18 7h-1.3A8 8 0 1 0 4 15.3"/></svg></span><span>Upload Vehicle Spreadsheet</span></div>
        <div class="pas-nav-row"><span class="pas-nav-icon"><svg viewBox="0 0 24 24"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><path d="M14 2v6h6"/><path d="M9 13h6"/><path d="M9 17h6"/></svg></span><span>Upload Vehicle Hire Invoice PDF</span></div>
        <div class="pas-nav-row"><span class="pas-nav-icon"><svg viewBox="0 0 24 24"><path d="M5 3l14 9-14 9V3z"/></svg></span><span>Run Reconciliation</span></div>
        <div class="pas-nav-row"><span class="pas-nav-icon"><svg viewBox="0 0 24 24"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><path d="M7 10l5 5 5-5"/><path d="M12 15V3"/></svg></span><span>Download Reconciliation<br>PDF</span></div>
        <div class="pas-nav-row"><span class="pas-nav-icon"><svg viewBox="0 0 24 24"><circle cx="11" cy="11" r="8"/><path d="M21 21l-4.3-4.3"/></svg></span><span>Smoke Crack</span></div>
        <div class="pas-sidebar-rule"></div>
        <div class="pas-sidebar-footer">PAS NW Ltd • v1.0.1 Vehicle Hire Simplified</div>
        """,
        unsafe_allow_html=True,
    )

st.markdown(
    """
    <div class="pas-hero">
      <div class="pas-hero-logo">PAS</div>
      <div class="pas-hero-text">PAS NW Ltd<span class="pas-hero-dot">•</span><span class="pas-hero-version">v1.0.1 Vehicle Hire Simplified</span></div>
    </div>
    """,
    unsafe_allow_html=True,
)



def render_bottom_chase():
    """Small non-intrusive PAS dump truck chase animation pinned to the bottom of the white content area."""
    st.markdown(
        """
        <div class="pas-bottom-chase-wrap" aria-hidden="true">
            <div class="pas-bottom-ground"></div>
            <div class="pas-chase-pack">
                <div class="pas-speed-lines"><span></span><span></span><span></span></div>
                <div class="pas-dust"><span></span><span></span><span></span></div>
                <div class="pas-truck-mini">
                    <div class="pas-truck-bed"></div>
                    <div class="pas-truck-logo">PAS</div>
                    <div class="pas-truck-cab"></div>
                    <div class="pas-truck-window"></div>
                    <div class="pas-truck-nose"></div>
                    <div class="pas-wheel back"></div>
                    <div class="pas-wheel front"></div>
                </div>
                <div class="pas-stickman">
                    <div class="pas-stick-head"></div>
                    <div class="pas-stick-body"></div>
                    <div class="pas-stick-arm-a"></div>
                    <div class="pas-stick-arm-b"></div>
                    <div class="pas-stick-leg-a"></div>
                    <div class="pas-stick-leg-b"></div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )




st.markdown(
    """
    <style>
    div[data-testid="stAlert"], div[data-testid="stAlert"] * {
        color: #0A0A0A !important;
    }
    div[data-testid="stAlert"] {
        border: 1px solid #e2ba00 !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# ===== Vehicle hire invoice reconciliation logic =====
try:
    from pypdf import PdfWriter
except Exception:
    try:
        from PyPDF2 import PdfWriter
    except Exception:
        PdfWriter = None

try:
    import fitz  # PyMuPDF
except Exception:
    fitz = None

try:
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except Exception:
    PatternFill = Font = Alignment = Border = Side = get_column_letter = None

VEHICLE_RESULT_COLUMNS = [
    "Component No",
    "Vehicle Registration",
    "Driver / Assigned User",
    "Job Number / Site",
    "On/Off Hire",
    "Invoice line value",
    "Total per vehicle",
    "Status",
]

DETAIL_COLUMNS = [
    "Component No",
    "Vehicle Registration",
    "Driver / Assigned User",
    "Job Number / Site",
    "On/Off Hire",
    "From",
    "To",
    "Invoice line value",
    "VAT",
    "Gross",
    "Status",
]


def render_selected_file_card(uploaded_file, file_kind="excel"):
    size = getattr(uploaded_file, "size", 0) or 0
    if size >= 1024 * 1024:
        size_text = f"{size / (1024 * 1024):.1f} MB"
    else:
        size_text = f"{size / 1024:.0f} KB"
    icon = "XLS" if file_kind == "excel" else "PDF"
    icon_class = "excel" if file_kind == "excel" else "pdf"
    st.markdown(
        f'''
        <div class="pas-file-card">
            <div class="pas-file-icon {icon_class}">{icon}</div>
            <div class="pas-file-main">
                <div class="pas-file-name">{escape(getattr(uploaded_file, "name", "Uploaded file"))}</div>
                <div class="pas-file-size">{size_text}</div>
            </div>
            <div class="pas-file-check">✓</div>
        </div>
        ''',
        unsafe_allow_html=True,
    )


def clean_cell(value) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "nat"}:
        return ""
    return text


def normalise_reg(value) -> str:
    return re.sub(r"[^A-Z0-9]+", "", clean_cell(value).upper())


def find_col(columns: List[str], keywords: List[str]) -> Optional[str]:
    norm_cols = {c: re.sub(r"[^a-z0-9]+", "", str(c).lower()) for c in columns}
    for key in keywords:
        nkey = re.sub(r"[^a-z0-9]+", "", key.lower())
        for col, ncol in norm_cols.items():
            if nkey == ncol:
                return col
    for key in keywords:
        nkey = re.sub(r"[^a-z0-9]+", "", key.lower())
        for col, ncol in norm_cols.items():
            if nkey in ncol:
                return col
    return None


def money_to_float(value) -> float:
    if value is None or pd.isna(value):
        return 0.0
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = str(value).replace(",", "")
    matches = re.findall(r"-?£?\s*([0-9]+(?:\.[0-9]{2})?)", text)
    if not matches:
        return 0.0
    try:
        return round(float(matches[-1]), 2)
    except Exception:
        return 0.0


def fmt_money(value) -> str:
    try:
        return f"£{float(value):,.2f}"
    except Exception:
        return "£0.00"



def load_vehicle_database(uploaded_file) -> Dict[str, Dict[str, str]]:
    """Load Vehicles.xlsx and create a registration lookup.

    Vehicle registration is matched by the Vehicle Reg/Registration column.
    Job Number / Site is forced from Column M (zero-based index 12), because
    that is the PAS job column even if the header changes.

    Off Hire - Sold can contain the same registration more than once, so the
    latest Date Returned row wins for that sheet.
    """
    xls = pd.ExcelFile(uploaded_file)
    sheet_lookup = {s.lower().strip(): s for s in xls.sheet_names}

    active_sheet = sheet_lookup.get("vehicles")
    offhire_sheet = sheet_lookup.get("off hire - sold") or sheet_lookup.get("off hire sold")

    rows_by_reg: Dict[str, List[Dict[str, str]]] = {}

    def add_sheet_rows(sheet_name: str, on_off_hire: str):
        if not sheet_name:
            return
        df = pd.read_excel(xls, sheet_name=sheet_name).dropna(how="all")
        cols = list(df.columns)

        # PSD invoice matching is still by vehicle registration.
        reg_col = find_col(cols, ["Vehicle Reg", "Vehicle Registration", "Registration", "Reg", "VRN"])
        if not reg_col and len(cols) >= 3:
            reg_col = cols[2]  # PAS Vehicles.xlsx normally stores reg in Column C.

        # PAS job/site is Column M, regardless of header wording.
        job_col = cols[12] if len(cols) > 12 else find_col(cols, ["Job Number / Site", "Job No", "Job", "Site", "Project", "Location"])

        driver_col = find_col(cols, ["Driver", "Driver Name", "Assigned User", "User", "Employee"])
        returned_col = find_col(cols, ["Date returned", "Date Returned", "Returned Date", "Off Hire Date"])
        supplier_col = find_col(cols, ["Supplier"])

        if not reg_col:
            return

        for _, row in df.iterrows():
            raw_reg = clean_cell(row.get(reg_col, ""))
            reg = normalise_reg(raw_reg)
            if not reg:
                continue

            returned_value = row.get(returned_col, "") if returned_col else ""
            returned_date = pd.to_datetime(returned_value, errors="coerce", dayfirst=True)

            rows_by_reg.setdefault(reg, []).append({
                "Vehicle Registration": raw_reg.upper(),
                "Driver / Assigned User": clean_cell(row.get(driver_col, "")) if driver_col else "",
                "Job Number / Site": clean_cell(row.get(job_col, "")) if job_col else "",
                "Supplier": clean_cell(row.get(supplier_col, "")) if supplier_col else "",
                "Source Sheet": sheet_name,
                "On/Off Hire": on_off_hire,
                "Date Returned": returned_date,
            })

    add_sheet_rows(active_sheet, "On Hire")
    add_sheet_rows(offhire_sheet, "Off Hire")

    lookup: Dict[str, Dict[str, str]] = {}
    for reg, rows in rows_by_reg.items():
        active_rows = [r for r in rows if r.get("On/Off Hire") == "On Hire"]
        offhire_rows = [r for r in rows if r.get("On/Off Hire") == "Off Hire"]

        if active_rows:
            chosen = active_rows[0]
        elif offhire_rows:
            # Latest returned record wins, fixing repeated off-hire registrations like CK75ENU.
            offhire_rows.sort(
                key=lambda r: pd.Timestamp.min if pd.isna(r.get("Date Returned")) else r.get("Date Returned"),
                reverse=True,
            )
            chosen = offhire_rows[0]
        else:
            chosen = rows[0]

        chosen = dict(chosen)
        chosen["Date Returned"] = "" if pd.isna(chosen.get("Date Returned")) else chosen.get("Date Returned").strftime("%d/%m/%Y")
        lookup[reg] = chosen

    return lookup


def read_pdf_pages(uploaded_pdf) -> Tuple[List[str], bytes]:
    pdf_bytes = uploaded_pdf.read()
    uploaded_pdf.seek(0)
    if PdfReader is None:
        raise RuntimeError("PDF reader is unavailable. Add pypdf to requirements.txt.")
    reader = PdfReader(io.BytesIO(pdf_bytes))
    pages = []
    for page in reader.pages:
        try:
            pages.append(page.extract_text() or "")
        except Exception:
            pages.append("")
    return pages, pdf_bytes


def extract_vehicle_hire_lines_from_text(pages: List[str]) -> List[Dict[str, str]]:
    """Extract PSD/Prohire vehicle component lines from a consolidated invoice."""
    rows: List[Dict[str, str]] = []
    seen = set()
    component_re = re.compile(
        r"(?P<component>[A-Z]\d{5,}/\d+)\s+"
        r"(?P<reg>[A-Z]{2}\d{2}[A-Z]{3})\s+"
        r"(?P<body>.*?)"
        r"(?P<from>\d{2}/\d{2}/\d{4})\s+"
        r"(?P<to>\d{2}/\d{2}/\d{4})\s+"
        r"(?:(?:\d{1,2}:\d{2}:\d{2})\s+)?"
        r"£(?P<net>[\d,]+\.\d{2})\s+£(?P<vat>[\d,]+\.\d{2})\s+£(?P<gross>[\d,]+\.\d{2})",
        re.I | re.S,
    )
    for page_no, text in enumerate(pages, start=1):
        flat = re.sub(r"\s+", " ", text or " ").strip()
        for match in component_re.finditer(flat):
            component = match.group("component").upper()
            if component in seen:
                continue
            seen.add(component)
            rows.append({
                "Component No": component,
                "Vehicle Registration": match.group("reg").upper(),
                "From": match.group("from"),
                "To": match.group("to"),
                "Invoice line value": money_to_float(match.group("net")),
                "VAT": money_to_float(match.group("vat")),
                "Gross": money_to_float(match.group("gross")),
                "Page": page_no,
            })
    return rows



def reconcile_vehicle_lines(lines: List[Dict[str, str]], vehicle_lookup: Dict[str, Dict[str, str]]) -> pd.DataFrame:
    out = []
    for line in lines:
        reg = normalise_reg(line.get("Vehicle Registration", ""))
        match = vehicle_lookup.get(reg, {})
        matched = bool(match)
        row = dict(line)
        row["Driver / Assigned User"] = match.get("Driver / Assigned User", "") if matched else ""
        row["Job Number / Site"] = match.get("Job Number / Site", "") if matched else ""
        row["On/Off Hire"] = match.get("On/Off Hire", "") if matched else ""
        row["Status"] = "Matched" if matched else "Unmatched"
        out.append(row)
    df = pd.DataFrame(out)
    if df.empty:
        return df

    vehicle_totals = df.groupby("Vehicle Registration", dropna=False)["Invoice line value"].sum().to_dict()
    df["Total per vehicle"] = df["Vehicle Registration"].map(vehicle_totals).fillna(df["Invoice line value"])
    return df



def make_vehicle_excel(summary_df: pd.DataFrame, detail_df: pd.DataFrame) -> bytes:
    """Create the simplified PAS vehicle hire Excel report.

    Tabs:
    - Vehicle Lines, with a Grand Total row at the bottom.
    - By Job.
    """
    output = io.BytesIO()
    detail = detail_df.copy()
    if detail.empty:
        detail = pd.DataFrame(columns=DETAIL_COLUMNS)

    for col in DETAIL_COLUMNS:
        if col not in detail.columns:
            detail[col] = ""

    vehicle_lines = detail[DETAIL_COLUMNS].copy()
    if not vehicle_lines.empty:
        grand_row = {col: "" for col in vehicle_lines.columns}
        grand_row["Job Number / Site"] = "Grand Total"
        grand_row["Invoice line value"] = float(vehicle_lines["Invoice line value"].sum())
        grand_row["VAT"] = float(vehicle_lines["VAT"].sum()) if "VAT" in vehicle_lines else 0
        grand_row["Gross"] = float(vehicle_lines["Gross"].sum()) if "Gross" in vehicle_lines else 0
        vehicle_lines = pd.concat([vehicle_lines, pd.DataFrame([grand_row])], ignore_index=True)

    by_job = (
        detail.groupby(["Job Number / Site"], dropna=False, as_index=False)
        .agg(**{
            "Total per job": ("Invoice line value", "sum"),
            "Vehicle lines": ("Component No", "count"),
        })
        .sort_values(["Job Number / Site"], kind="stable")
    )

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        vehicle_lines.to_excel(writer, index=False, sheet_name="Vehicle Lines")
        by_job.to_excel(writer, index=False, sheet_name="By Job")

        wb = writer.book
        if PatternFill and Font and Alignment and Border and Side and get_column_letter:
            yellow = PatternFill("solid", fgColor="FFD400")
            header_font = Font(bold=True, color="000000")
            total_fill = PatternFill("solid", fgColor="FFF2CC")
            total_font = Font(bold=True, color="000000")
            thin = Side(style="thin", color="D9D9D9")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for ws in wb.worksheets:
                ws.freeze_panes = "A2"
                for cell in ws[1]:
                    cell.fill = yellow
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        cell.border = border
                # Highlight the grand total row on Vehicle Lines.
                if ws.title == "Vehicle Lines" and ws.max_row >= 2:
                    label_col = None
                    for cell in ws[1]:
                        if cell.value == "Job Number / Site":
                            label_col = cell.column
                            break
                    if label_col:
                        for row_idx in range(2, ws.max_row + 1):
                            if ws.cell(row_idx, label_col).value == "Grand Total":
                                for cell in ws[row_idx]:
                                    cell.fill = total_fill
                                    cell.font = total_font
                for col_idx, column_cells in enumerate(ws.columns, start=1):
                    max_len = 12
                    for cell in column_cells:
                        max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 45)
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        header = str(ws.cell(1, cell.column).value).lower()
                        if isinstance(cell.value, (int, float)) and any(word in header for word in ["value", "total", "vat", "gross"]):
                            cell.number_format = '£#,##0.00'
    output.seek(0)
    return output.getvalue()



def make_annotated_vehicle_pdf(original_pdf_bytes: bytes, detail_df: pd.DataFrame) -> bytes:
    """Annotate each matched PSD vehicle line with Job Number / Site.

    The previous version widened the page before reading text positions, which
    could stop anchors being found on some PDFs. This version finds the vehicle
    registration coordinates first, then adds the side-note margin and writes
    the job/site note on the same row.
    """
    if fitz is None or detail_df is None or detail_df.empty:
        return original_pdf_bytes
    try:
        doc = fitz.open(stream=original_pdf_bytes, filetype="pdf")
        margin_width = 135

        for page_index in range(len(doc)):
            page = doc[page_index]
            page_rows = detail_df[detail_df["Page"] == page_index + 1].copy()
            if page_rows.empty:
                continue

            old_width = page.rect.width
            old_height = page.rect.height

            # Find text anchors before resizing the page.
            words = sorted(page.get_text("words"), key=lambda w: (round(w[1], 1), w[0]))
            anchors = []
            for w in words:
                word = normalise_reg(w[4])
                if re.fullmatch(r"[A-Z]{2}\d{2}[A-Z]{3}", word):
                    anchors.append({
                        "reg": word,
                        "y": (float(w[1]) + float(w[3])) / 2,
                        "used": False,
                    })

            new_width = old_width + margin_width
            page.set_mediabox(fitz.Rect(0, 0, new_width, old_height))
            page.draw_rect(fitz.Rect(old_width, 0, new_width, old_height), color=None, fill=(1, 1, 1), overlay=True)
            page.draw_line((old_width + 2, 0), (old_width + 2, old_height), color=(0.84, 0.84, 0.84), width=0.5, overlay=True)

            for _, row in page_rows.iterrows():
                site = clean_cell(row.get("Job Number / Site", ""))
                if not site:
                    continue
                reg = normalise_reg(row.get("Vehicle Registration", ""))
                chosen = None
                for anchor in anchors:
                    if not anchor["used"] and anchor["reg"] == reg:
                        chosen = anchor
                        break
                if chosen is None:
                    # Fallback: use the line order from extraction if the text anchor cannot be found.
                    row_number = list(page_rows.index).index(row.name)
                    approx_top = 124
                    approx_step = 17
                    chosen = {"y": approx_top + row_number * approx_step, "used": True}
                else:
                    chosen["used"] = True

                driver = clean_cell(row.get("Driver / Assigned User", ""))
                note = site if not driver else f"{site} | {driver}"
                y = float(chosen["y"]) + 3.0
                text_box = fitz.Rect(old_width + 8, y - 8, new_width - 5, y + 13)
                page.insert_textbox(
                    text_box,
                    note[:58],
                    fontsize=7.0,
                    fontname="helv",
                    color=(0, 0, 0),
                    align=fitz.TEXT_ALIGN_LEFT,
                    overlay=True,
                )

        out = io.BytesIO()
        doc.save(out, garbage=4, deflate=True)
        doc.close()
        return out.getvalue()
    except Exception:
        return original_pdf_bytes


def render_results_table(df: pd.DataFrame):
    if df is None or df.empty:
        st.markdown('<div class="pas-unmatched-pill">Results</div>', unsafe_allow_html=True)
        st.markdown('<div class="pas-table-wrap"><table class="pas-table"><tbody><tr><td>No vehicle lines found.</td></tr></tbody></table></div>', unsafe_allow_html=True)
        return
    display_df = df.copy()
    for col in VEHICLE_RESULT_COLUMNS:
        if col not in display_df.columns:
            display_df[col] = ""
    display_df = display_df[VEHICLE_RESULT_COLUMNS]
    money_cols = {"Invoice line value", "Total per vehicle"}
    rows_html = []
    for _, row in display_df.iterrows():
        cells = []
        for col in display_df.columns:
            value = row.get(col, "")
            text = fmt_money(value) if col in money_cols else clean_cell(value)
            cells.append(f"<td>{escape(text)}</td>")
        rows_html.append(f"<tr>{''.join(cells)}</tr>")
    header_html = "".join(f"<th>{escape(col)}</th>" for col in display_df.columns)
    st.markdown('<div class="pas-unmatched-pill">Vehicle Hire Lines</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="pas-table-wrap"><table class="pas-table"><thead><tr>{header_html}</tr></thead><tbody>{"".join(rows_html)}</tbody></table></div>', unsafe_allow_html=True)


up_col1, up_col2 = st.columns(2)
with up_col1:
    st.markdown('<div class="pas-upload-card"><div class="pas-upload-title">Upload Vehicles.xlsx</div>', unsafe_allow_html=True)
    vehicles_file = st.file_uploader("Upload Vehicles.xlsx", type=["xlsx", "xls"], label_visibility="collapsed", key="vehicles_upload")
    if vehicles_file:
        render_selected_file_card(vehicles_file, "excel")
    st.markdown('</div>', unsafe_allow_html=True)
with up_col2:
    st.markdown('<div class="pas-upload-card"><div class="pas-upload-title">Upload Vehicle Hire Invoice PDF</div>', unsafe_allow_html=True)
    hire_pdf = st.file_uploader("Upload Vehicle Hire Invoice PDF", type=["pdf"], label_visibility="collapsed", key="vehicle_hire_pdf_upload")
    if hire_pdf:
        render_selected_file_card(hire_pdf, "pdf")
    st.markdown('</div>', unsafe_allow_html=True)

run = st.button("▶  Run reconciliation", use_container_width=True)

if "vehicle_hire_reconciliation_results" not in st.session_state:
    st.session_state["vehicle_hire_reconciliation_results"] = None

if run:
    if not vehicles_file or not hire_pdf:
        st.warning("Please upload both Vehicles.xlsx and the vehicle hire invoice PDF.")
        st.stop()
    try:
        with st.spinner("Reading vehicle database..."):
            vehicle_lookup = load_vehicle_database(vehicles_file)
        with st.spinner("Reading vehicle hire invoice PDF..."):
            pages, original_pdf_bytes = read_pdf_pages(hire_pdf)
            lines = extract_vehicle_hire_lines_from_text(pages)
        with st.spinner("Reconciling vehicle hire lines..."):
            all_df = reconcile_vehicle_lines(lines, vehicle_lookup)
            if all_df.empty:
                st.warning("No vehicle hire lines could be extracted from the PDF.")
                st.stop()

        total = len(all_df)
        matched = int((all_df["Status"] == "Matched").sum())
        unmatched = int((all_df["Status"] == "Unmatched").sum())
        match_pct = round((matched / total) * 100, 1) if total else 0.0
        invoice_total = round(float(all_df["Invoice line value"].sum()), 2)

        by_job = all_df.groupby("Job Number / Site", dropna=False)["Invoice line value"].sum().reset_index()
        by_job["Invoice line value"] = by_job["Invoice line value"].round(2)

        summary_df = pd.DataFrame({
            "Metric": [
                "Total vehicle lines",
                "Matched vehicles",
                "Unmatched vehicles",
                "Match percentage",
                "Invoice total",
                "Run date/time",
            ],
            "Value": [
                total,
                matched,
                unmatched,
                f"{match_pct}%",
                invoice_total,
                datetime.now().strftime("%d/%m/%Y %H:%M"),
            ],
        })

        excel_bytes = make_vehicle_excel(summary_df, all_df)
        annotated_pdf_bytes = make_annotated_vehicle_pdf(original_pdf_bytes, all_df)

        stamp = datetime.now().strftime("%Y%m%d_%H%M")
        st.session_state["vehicle_hire_reconciliation_results"] = {
            "all_df": all_df,
            "summary_df": summary_df,
            "by_job": by_job,
            "excel_bytes": excel_bytes,
            "annotated_pdf_bytes": annotated_pdf_bytes,
            "total": total,
            "matched": matched,
            "unmatched": unmatched,
            "match_pct": match_pct,
            "invoice_total": invoice_total,
            "excel_filename": f"PAS_Vehicle_Hire_Reconciliation_{stamp}.xlsx",
            "pdf_filename": f"PAS_Vehicle_Hire_Invoice_Annotated_{stamp}.pdf",
        }
    except Exception as e:
        st.error(f"Something went wrong: {e}")
        st.exception(e)

results = st.session_state.get("vehicle_hire_reconciliation_results")

if results is not None:
    total = results["total"]
    matched = results["matched"]
    unmatched = results["unmatched"]
    match_pct = results["match_pct"]
    invoice_total = results["invoice_total"]
    all_df = results["all_df"]

    c1, c2, c3, c4, c5 = st.columns(5)
    with c1:
        st.markdown(f'<div class="kpi-card"><div class="kpi-icon"><svg viewBox="0 0 24 24"><path d="M8 7V3h8l4 4v14H6V7z"/><path d="M16 3v5h5"/><path d="M9 13h6"/><path d="M9 17h4"/><path d="M4 7h2v14h12"/></svg></div><div><div class="kpi-label">Total vehicle lines</div><div class="kpi-value">{total}</div><div class="kpi-sub">Detected lines</div></div></div>', unsafe_allow_html=True)
    with c2:
        st.markdown(f'<div class="kpi-card kpi-matched"><div class="kpi-icon"><svg viewBox="0 0 24 24"><circle cx="12" cy="12" r="9"/><path d="M8 12.5l2.7 2.7L16.5 9"/></svg></div><div><div class="kpi-label">Matched vehicles</div><div class="kpi-value">{matched}</div><div class="kpi-sub">Matched by reg</div></div></div>', unsafe_allow_html=True)
    with c3:
        st.markdown(f'<div class="kpi-card kpi-unmatched"><div class="kpi-icon"><svg viewBox="0 0 24 24"><path d="M12 3l10 18H2L12 3z"/><path d="M12 9v5"/><path d="M12 18h.01"/></svg></div><div><div class="kpi-label">Unmatched vehicles</div><div class="kpi-value">{unmatched}</div><div class="kpi-sub">Blank job/site</div></div></div>', unsafe_allow_html=True)
    with c4:
        st.markdown(f'<div class="kpi-card"><div class="kpi-icon"><svg viewBox="0 0 24 24"><path d="M3 20h18"/><path d="M6 16v-4"/><path d="M11 16V8"/><path d="M16 16v-6"/><path d="M19 6l-5 5-3-3-5 5"/></svg></div><div><div class="kpi-label">Match %</div><div class="kpi-value">{match_pct}%</div><div class="kpi-sub">Core KPI</div></div></div>', unsafe_allow_html=True)
    with c5:
        st.markdown(f'<div class="kpi-card"><div class="kpi-icon"><svg viewBox="0 0 24 24"><path d="M4 19h16"/><path d="M7 19V5h10v14"/><path d="M9 9h6"/><path d="M9 13h6"/></svg></div><div><div class="kpi-label">Invoice total</div><div class="kpi-value">{fmt_money(invoice_total)}</div><div class="kpi-sub">Net value</div></div></div>', unsafe_allow_html=True)

    st.markdown('<div class="pas-results-title">Results</div>', unsafe_allow_html=True)
    render_results_table(all_df)

    dl_left, dl_mid, dl_right = st.columns([1, 1, 1])
    with dl_left:
        st.download_button(
            "⬇  Download Excel report",
            data=results["excel_bytes"],
            file_name=results["excel_filename"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with dl_right:
        st.download_button(
            "⬇  Download annotated PDF",
            data=results["annotated_pdf_bytes"],
            file_name=results["pdf_filename"],
            mime="application/pdf",
            use_container_width=True,
        )
else:
    st.info("Upload Vehicles.xlsx and the vehicle hire invoice PDF, then click Run reconciliation.")
    render_bottom_chase()
